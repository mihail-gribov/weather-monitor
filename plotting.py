#!/usr/bin/env python3
"""
Weather data plotting module.

This module provides functionality for creating charts and graphs
from weather data stored in the database.
"""

import sqlite3
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
import yaml



try:
    import plotext as plt
    PLOTEXT_AVAILABLE = True
except ImportError:
    PLOTEXT_AVAILABLE = False
    logging.warning("plotext not available, ASCII plotting disabled")

try:
    import matplotlib.pyplot as mplt
    import matplotlib.dates as mdates
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    logging.warning("matplotlib not available, file plotting disabled")


class WeatherPlotter:
    """Weather data plotting and visualization."""
    
    def __init__(self, db_path: str, config: Dict[str, Any]):
        """Initialize plotter with database path and configuration."""
        self.db_path = db_path
        self.config = config.get('plotting', {})
        self.default_colors = self.config.get('default_colors', 
                                            ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd'])
        self.ascii_symbols = self.config.get('ascii_symbols', ['─', '┄', '┅', '┈', '┉'])
        self.figure_size = self.config.get('figure_size', [10, 6])
        self.dpi = self.config.get('dpi', 100)
        
        # Metric mapping
        self.metric_mapping = {
            'temperature': 'temperature',
            'humidity': 'humidity',
            'pressure': 'pressure',
            'wind-speed': 'wind_speed',
            'precipitation': 'precipitation'
        }
        
        # Available metrics
        self.available_metrics = list(self.metric_mapping.keys())
        
        # Export formats
        self.supported_chart_formats = ['png', 'svg', 'pdf']
        self.supported_data_formats = ['csv', 'json', 'excel']
    
    def get_region_color(self, region_code: str) -> str:
        """Generate stable color for a specific region based on its position."""
        # For plotting, we'll use a simple approach with predefined colors
        # This ensures consistent colors for static plots
        hash_value = hash(region_code)
        color_index = abs(hash_value) % len(self.default_colors)
        return self.default_colors[color_index]
    
    def get_colors_for_regions(self, region_codes: List[str]) -> Dict[str, str]:
        """Get stable colors for multiple regions."""
        colors = {}
        for region_code in region_codes:
            colors[region_code] = self.get_region_color(region_code)
        return colors
    
    def get_metric_data(self, metric: str, regions: List[str], hours: int) -> Dict[str, List[Tuple[datetime, float]]]:
        """Get weather data for specified metric and regions."""
        if metric not in self.metric_mapping:
            raise ValueError(f"Unknown metric: {metric}. Available: {self.available_metrics}")
        
        db_column = self.metric_mapping[metric]
        cutoff_time = datetime.now() - timedelta(hours=hours)
        
        conn = sqlite3.connect(self.db_path, detect_types=sqlite3.PARSE_DECLTYPES)
        try:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT region_code, timestamp, {} 
                FROM weather_data 
                WHERE region_code IN ({}) 
                AND timestamp >= ? 
                AND {} IS NOT NULL
                ORDER BY timestamp
            '''.format(db_column, ','.join(['?'] * len(regions)), db_column), 
            regions + [cutoff_time.strftime('%Y-%m-%d %H:%M:%S')])
            
            data = {}
            for row in cursor.fetchall():
                region_code, timestamp_str, value = row
                if region_code not in data:
                    data[region_code] = []
                # Handle both ISO format and SQLite datetime format
                try:
                    timestamp = datetime.fromisoformat(timestamp_str)
                except ValueError:
                    # Try parsing SQLite datetime format
                    timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                data[region_code].append((timestamp, value))
            
            return data
        finally:
            conn.close()
    
    def validate_regions(self, regions: List[str]) -> List[str]:
        """Validate that specified regions exist in database."""
        conn = sqlite3.connect(self.db_path, detect_types=sqlite3.PARSE_DECLTYPES)
        try:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT DISTINCT region_code 
                FROM weather_data 
                WHERE region_code IN ({})
            '''.format(','.join(['?'] * len(regions))), regions)
            
            existing_regions = [row[0] for row in cursor.fetchall()]
            missing_regions = set(regions) - set(existing_regions)
            
            if missing_regions:
                logging.warning(f"Missing regions: {missing_regions}")
            
            return existing_regions
        finally:
            conn.close()
    
    def get_available_metrics(self) -> List[str]:
        """Get list of available metrics."""
        return self.available_metrics.copy()
    
    def get_available_regions(self) -> List[str]:
        """Get list of available regions from database."""
        conn = sqlite3.connect(self.db_path, detect_types=sqlite3.PARSE_DECLTYPES)
        try:
            cursor = conn.cursor()
            cursor.execute('SELECT DISTINCT region_code FROM weather_data ORDER BY region_code')
            regions = [row[0] for row in cursor.fetchall()]
            return regions
        finally:
            conn.close()
    
    def plot_metric_ascii(self, metric: str, regions: List[str], hours: int) -> str:
        """Create ASCII plot for specified metric and regions."""
        if not PLOTEXT_AVAILABLE:
            raise ImportError("plotext not available for ASCII plotting")
        
        # Validate regions
        valid_regions = self.validate_regions(regions)
        if not valid_regions:
            return "No valid regions found for plotting"
        
        # Get data
        data = self.get_metric_data(metric, regions, hours)
        if not data:
            return f"No data found for {metric} in the last {hours} hours"
        
        # Clear previous plots
        plt.clear_figure()
        
        # Set up the plot
        plt.plot_size(80, 20)
        plt.title(f"{metric.title()} - Last {hours} hours")
        
        # Plot data for each region
        for i, (region_code, region_data) in enumerate(data.items()):
            if not region_data:
                continue
                
            timestamps = [point[0] for point in region_data]
            values = [point[1] for point in region_data]
            
            # Convert timestamps to hours from start
            start_time = min(timestamps)
            hours_from_start = [(t - start_time).total_seconds() / 3600 for t in timestamps]
            
            # Plot with different symbols for each region
            symbol = self.ascii_symbols[i % len(self.ascii_symbols)]
            plt.plot(hours_from_start, values, label=region_code, marker=symbol)
        
        # Set labels and legend
        plt.xlabel("Hours")
        plt.ylabel(metric.title())
        
        # Show the plot and return empty string since show() outputs directly
        plt.show()
        return ""
    
    def plot_metric_file(self, metric: str, regions: List[str], hours: int, filename: str) -> bool:
        """Create file plot for specified metric and regions."""
        if not MATPLOTLIB_AVAILABLE:
            raise ImportError("matplotlib not available for file plotting")
        
        # Validate regions
        valid_regions = self.validate_regions(regions)
        if not valid_regions:
            logging.error("No valid regions found for plotting")
            return False
        
        # Get data
        data = self.get_metric_data(metric, regions, hours)
        if not data:
            logging.error(f"No data found for {metric} in the last {hours} hours")
            return False
        
        try:
            # Create figure
            fig, ax = mplt.subplots(figsize=tuple(self.figure_size), dpi=self.dpi)
            
            # Plot data for each region
            for region_code, region_data in data.items():
                if not region_data:
                    continue
                    
                timestamps = [point[0] for point in region_data]
                values = [point[1] for point in region_data]
                
                # Use stable color for region
                color = self.get_region_color(region_code)
                ax.plot(timestamps, values, label=region_code, color=color, linewidth=2)
            
            # Customize the plot
            ax.set_title(f"{metric.title()} - Last {hours} hours", fontsize=14, fontweight='bold')
            ax.set_xlabel("Time", fontsize=12)
            ax.set_ylabel(metric.title(), fontsize=12)
            ax.grid(True, alpha=0.3)
            ax.legend()
            
            # Format x-axis dates
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
            ax.xaxis.set_major_locator(mdates.HourLocator(interval=max(1, hours // 6)))
            mplt.setp(ax.xaxis.get_majorticklabels(), rotation=45)
            
            # Adjust layout and save
            fig.tight_layout()
            fig.savefig(filename, dpi=self.dpi, bbox_inches='tight')
            mplt.close(fig)
            
            logging.info(f"Plot saved to {filename}")
            return True
            
        except Exception as e:
            logging.error(f"Error creating plot: {e}")
            return False

    def export_chart_data(self, metric: str, regions: List[str], hours: int, filename: str, format: str = 'csv') -> bool:
        """Export chart data to various formats."""
        if format not in self.supported_data_formats:
            raise ValueError(f"Unsupported format: {format}. Supported: {self.supported_data_formats}")
        
        # Validate regions
        valid_regions = self.validate_regions(regions)
        if not valid_regions:
            logging.error("No valid regions found for export")
            return False
        
        # Get data
        data = self.get_metric_data(metric, regions, hours)
        if not data:
            logging.error(f"No data found for {metric} in the last {hours} hours")
            return False
        
        try:
            if format == 'csv':
                return self._export_to_csv(data, metric, filename)
            elif format == 'json':
                return self._export_to_json(data, metric, filename)
            elif format == 'excel':
                return self._export_to_excel(data, metric, filename)
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            logging.error(f"Error exporting data: {e}")
            return False

    def _export_to_csv(self, data: Dict[str, List[Tuple[datetime, float]]], metric: str, filename: str) -> bool:
        """Export data to CSV format."""
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow(['Region', 'Timestamp', metric.title()])
            
            # Write data
            for region_code, region_data in data.items():
                for timestamp, value in region_data:
                    writer.writerow([region_code, timestamp.isoformat(), value])
        
        logging.info(f"Data exported to CSV: {filename}")
        return True

    def _export_to_json(self, data: Dict[str, List[Tuple[datetime, float]]], metric: str, filename: str) -> bool:
        """Export data to JSON format."""
        import json
        
        # Convert data to JSON-serializable format
        json_data = {
            'metric': metric,
            'export_time': datetime.now().isoformat(),
            'regions': {}
        }
        
        for region_code, region_data in data.items():
            json_data['regions'][region_code] = [
                {
                    'timestamp': timestamp.isoformat(),
                    'value': value
                }
                for timestamp, value in region_data
            ]
        
        with open(filename, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False)
        
        logging.info(f"Data exported to JSON: {filename}")
        return True

    def _export_to_excel(self, data: Dict[str, List[Tuple[datetime, float]]], metric: str, filename: str) -> bool:
        """Export data to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError("openpyxl not available for Excel export. Install with: pip install openpyxl")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{metric.title()} Data"
        
        # Write header
        header_font = Font(bold=True)
        ws['A1'] = 'Region'
        ws['B1'] = 'Timestamp'
        ws['C1'] = metric.title()
        
        for cell in ws[1]:
            cell.font = header_font
        
        # Write data
        row = 2
        for region_code, region_data in data.items():
            for timestamp, value in region_data:
                ws[f'A{row}'] = region_code
                ws[f'B{row}'] = timestamp.isoformat()
                ws[f'C{row}'] = value
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        wb.save(filename)
        logging.info(f"Data exported to Excel: {filename}")
        return True

    def generate_report(self, metric: str, regions: List[str], hours: int, filename: str) -> bool:
        """Generate PDF report with charts and statistics."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
        except ImportError:
            raise ImportError("reportlab not available for PDF reports. Install with: pip install reportlab")
        
        # Validate regions
        valid_regions = self.validate_regions(regions)
        if not valid_regions:
            logging.error("No valid regions found for report")
            return False
        
        # Get data
        data = self.get_metric_data(metric, regions, hours)
        if not data:
            logging.error(f"No data found for {metric} in the last {hours} hours")
            return False
        
        try:
            # Create PDF document
            doc = SimpleDocTemplate(filename, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30
            )
            story.append(Paragraph(f"Weather Data Report - {metric.title()}", title_style))
            story.append(Spacer(1, 12))
            
            # Summary
            story.append(Paragraph("Summary", styles['Heading2']))
            story.append(Paragraph(f"Metric: {metric.title()}", styles['Normal']))
            story.append(Paragraph(f"Time Range: Last {hours} hours", styles['Normal']))
            story.append(Paragraph(f"Regions: {', '.join(valid_regions)}", styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Statistics table
            story.append(Paragraph("Statistics", styles['Heading2']))
            
            # Calculate statistics
            stats_data = [['Region', 'Min', 'Max', 'Average', 'Data Points']]
            for region_code, region_data in data.items():
                if region_data:
                    values = [point[1] for point in region_data]
                    stats_data.append([
                        region_code,
                        f"{min(values):.2f}",
                        f"{max(values):.2f}",
                        f"{sum(values) / len(values):.2f}",
                        str(len(values))
                    ])
            
            # Create table
            table = Table(stats_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            story.append(Spacer(1, 12))
            
            # Export timestamp
            story.append(Paragraph(f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            
            # Build PDF
            doc.build(story)
            logging.info(f"PDF report generated: {filename}")
            return True
            
        except Exception as e:
            logging.error(f"Error generating PDF report: {e}")
            return False

    def get_supported_formats(self) -> Dict[str, List[str]]:
        """Get supported export formats."""
        return {
            'chart_formats': self.supported_chart_formats,
            'data_formats': self.supported_data_formats
        }


def main():
    """Main function for plotting module."""
    print("Weather Plotting Module")
    print("Available metrics:", WeatherPlotter("weather_data.db", {}).get_available_metrics())


if __name__ == "__main__":
    main()
